import sys
from openpyxl import load_workbook
import osbook = load_workbook("Book1.xlsx")
ws = book.worksheets[0]
count = 0
for cell in ws["C"]:
    if cell.value is None:
        break
    else:
        count = cell.row + 1
print(count)with open("string.txt", "r") as fi:
    variety_txt = []
    release = []
    maintainer = []
    string = ''
    for x in fi:
        for i in range(len(x)):
            if x[i] == '.' and x[i + 1] == ' ':
                string = (x[i + 2:])
                break
        for i in range(len(string)):
            if string[i] == ' ' and string[i + 1] != ' ':
                variety_txt.append(string)
                break
'''
for i in range(len(variety_txt)):
    print(variety_txt[i])
'''
def save(variety_par):
    for ix in range(len(variety_par)):
        if 'variety' in variety_par[ix]:
            ws.cell(row=count, column=1).value = variety_par[ix]
        if 'Year' in variety_par[ix]:
            ws.cell(row=count, column=2).value = variety_par[ix]
        if 'Breeder' in variety_par[ix]:
            ws.cell(row=count, column=3).value = variety_par[ix]
    book.save("Book1.xlsx")
    os.execv(sys.argv[0], sys.argv)
save(variety_txt)
